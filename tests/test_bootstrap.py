import sqlite3
from pathlib import Path

import openpyxl
import pytest

from jumia_feed_sync import bootstrap, db

_HEADER = ["Name", "SellerSKU", "ParentSKU", "Brand", "PrimaryCategory", "AdditionalCategory", "Price_KES"]


def _make_template(path: Path, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HEADER)
    for row in rows:
        ws.append([row.get(col) for col in _HEADER])
    wb.save(path)


@pytest.fixture
def conn():
    connection = sqlite3.connect(":memory:")
    db.migrate(connection)
    return connection


def test_harvest_extracts_brand_and_category_pairs(tmp_path):
    path = tmp_path / "template.xlsx"
    _make_template(
        path,
        [
            {"Name": "Widget A", "SellerSKU": "A1", "Brand": "1045133 - Generic",
             "PrimaryCategory": "1002708 - Computing / Printer Ink"},
            {"Name": "Widget B", "SellerSKU": "A2", "Brand": "1045133 - Generic",
             "PrimaryCategory": "1002708 - Computing / Printer Ink",
             "AdditionalCategory": "1002999 - Computing / Accessories"},
        ],
    )
    row_count, pairs = bootstrap.harvest_from_workbook(str(path))
    assert row_count == 2
    assert ("brand", "1045133", "Generic") in pairs
    assert ("category", "1002708", "Computing / Printer Ink") in pairs
    assert ("category", "1002999", "Computing / Accessories") in pairs


def test_harvest_skips_malformed_values(tmp_path):
    path = tmp_path / "template.xlsx"
    _make_template(path, [{"Name": "Widget C", "SellerSKU": "A3", "Brand": "Generic", "PrimaryCategory": "no id here"}])
    _, pairs = bootstrap.harvest_from_workbook(str(path))
    assert pairs == []


def test_harvest_skips_fully_blank_rows(tmp_path):
    path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HEADER)
    ws.append(["Widget A", "A1", None, "1045133 - Generic", "1002708 - Computing / Printer Ink", None, 1500])
    ws.append([None] * len(_HEADER))
    wb.save(path)

    row_count, pairs = bootstrap.harvest_from_workbook(str(path))
    assert row_count == 1
    assert len(pairs) == 2


def test_upsert_catalog_dedupes_and_tracks_new(conn):
    pairs = [
        ("brand", "1045133", "Generic"),
        ("brand", "1045133", "Generic"),
        ("category", "1002708", "Computing / Printer Ink"),
    ]
    found, new = bootstrap.upsert_catalog(conn, pairs)
    assert found == 2
    assert new == 2
    assert conn.execute("SELECT COUNT(*) FROM id_label_catalog").fetchone()[0] == 2


def test_upsert_catalog_second_run_reports_zero_new(conn):
    pairs = [("brand", "1045133", "Generic")]
    bootstrap.upsert_catalog(conn, pairs)
    found, new = bootstrap.upsert_catalog(conn, pairs)
    assert found == 1
    assert new == 0


def test_harvest_end_to_end(tmp_path, conn):
    path = tmp_path / "template.xlsx"
    _make_template(
        path,
        [{"Name": "Widget A", "SellerSKU": "A1", "Brand": "1045133 - Generic",
          "PrimaryCategory": "1002708 - Computing / Printer Ink"}],
    )
    summary = bootstrap.harvest(conn, str(path))
    assert summary.rows_scanned == 1
    assert summary.pairs_found == 2
    assert summary.pairs_new == 2


def _make_guidelines(path: Path, brands: list[str], categories: list[str]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    brands_ws = wb.create_sheet("Brands")
    brands_ws.append(["CODE - BRAND_SYSTEM_NAME"])
    for value in brands:
        brands_ws.append([value])
    categories_ws = wb.create_sheet("Categories")
    categories_ws.append(["CATEGORIES"])
    for value in categories:
        categories_ws.append([value])
    wb.save(path)


def test_harvest_from_guidelines_extracts_brands_and_categories(tmp_path):
    path = tmp_path / "guidelines.xlsx"
    _make_guidelines(
        path,
        brands=["1118344 - Ugreen", "1036890 - Epson"],
        categories=["1000055 - Computing / Computer Accessories / Audio & Video Accessories"],
    )
    row_count, pairs = bootstrap.harvest_from_guidelines(str(path))
    assert row_count == 3
    assert ("brand", "1118344", "Ugreen") in pairs
    assert ("brand", "1036890", "Epson") in pairs
    assert ("category", "1000055", "Computing / Computer Accessories / Audio & Video Accessories") in pairs


def test_harvest_guidelines_tags_source(tmp_path, conn):
    path = tmp_path / "guidelines.xlsx"
    _make_guidelines(path, brands=["1118344 - Ugreen"], categories=[])
    bootstrap.harvest_guidelines(conn, str(path))
    source = conn.execute(
        "SELECT source FROM id_label_catalog WHERE kind = 'brand' AND jumia_id = '1118344'"
    ).fetchone()[0]
    assert source == "jumia_reference"


def test_template_and_guidelines_harvests_coexist_in_catalog(tmp_path, conn):
    template_path = tmp_path / "template.xlsx"
    _make_template(
        template_path,
        [{"Name": "Widget A", "SellerSKU": "A1", "Brand": "1045133 - Generic",
          "PrimaryCategory": "1002708 - Computing / Printer Ink"}],
    )
    guidelines_path = tmp_path / "guidelines.xlsx"
    _make_guidelines(guidelines_path, brands=["1118344 - Ugreen"], categories=[])

    bootstrap.harvest(conn, str(template_path))
    bootstrap.harvest_guidelines(conn, str(guidelines_path))

    assert conn.execute("SELECT COUNT(*) FROM id_label_catalog").fetchone()[0] == 3
