import openpyxl
import pytest

from jumia_feed_sync.export import write_export, write_rejects_csv
from jumia_feed_sync.rules import Issue


def _make_template(path, header):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    wb.save(path)


def test_write_export_uses_template_header_order_not_dict_order(tmp_path):
    template_path = tmp_path / "template.xlsx"
    _make_template(template_path, ["SellerSKU", "Name", "Price_KES"])

    rows = [{"Price_KES": 1500, "Name": "Widget", "SellerSKU": "A1"}]
    out_path = tmp_path / "out.xlsx"
    written = write_export(rows, str(template_path), str(out_path))

    assert written == 1
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert [c.value for c in ws[1]] == ["SellerSKU", "Name", "Price_KES"]
    assert [c.value for c in ws[2]] == ["A1", "Widget", 1500]


def test_write_export_missing_field_becomes_blank_cell(tmp_path):
    template_path = tmp_path / "template.xlsx"
    _make_template(template_path, ["SellerSKU", "Brand"])

    out_path = tmp_path / "out.xlsx"
    write_export([{"SellerSKU": "A1"}], str(template_path), str(out_path))

    wb = openpyxl.load_workbook(out_path)
    assert [c.value for c in wb.active[2]] == ["A1", None]


def test_write_export_clears_existing_template_data_rows(tmp_path):
    template_path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SellerSKU", "Name"])
    ws.append(["OLD-1", "Old row that must not survive export"])
    wb.save(template_path)

    out_path = tmp_path / "out.xlsx"
    write_export([{"SellerSKU": "NEW-1", "Name": "New row"}], str(template_path), str(out_path))

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws.max_row == 2
    assert [c.value for c in ws[2]] == ["NEW-1", "New row"]


def test_write_export_header_row_untouched(tmp_path):
    template_path = tmp_path / "template.xlsx"
    _make_template(template_path, ["SellerSKU", "Name"])

    out_path = tmp_path / "out.xlsx"
    write_export([], str(template_path), str(out_path))

    wb = openpyxl.load_workbook(out_path)
    assert [c.value for c in wb.active[1]] == ["SellerSKU", "Name"]


@pytest.fixture
def issues():
    return [
        Issue(sku="A1", field="Name", severity="block", rule_id="name_length", message="too short"),
        Issue(sku="A2", field=None, severity="warn", rule_id="desc_not_title", message="duplicate"),
        Issue(sku="A3", field="Brand", severity="block", rule_id="brand_format", message="bad format"),
    ]


def test_write_rejects_csv_only_includes_blocked(tmp_path, issues):
    path = tmp_path / "rejects.csv"
    count = write_rejects_csv(issues, str(path))
    assert count == 2
    content = path.read_text(encoding="utf-8")
    assert "A1" in content and "A3" in content
    assert "A2" not in content


def test_write_rejects_csv_has_header(tmp_path, issues):
    path = tmp_path / "rejects.csv"
    write_rejects_csv(issues, str(path))
    first_line = path.read_text(encoding="utf-8").splitlines()[0]
    assert first_line == "sku,rule_id,field,message"
