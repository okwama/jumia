import sqlite3
from pathlib import Path

import pytest

from jumia_feed_sync import config, db, image, pipeline

REAL_RULES_PATH = Path(__file__).parent.parent / "config" / "rules.yaml"


@pytest.fixture
def conn():
    connection = sqlite3.connect(":memory:")
    db.migrate(connection)
    return connection


@pytest.fixture(autouse=True)
def no_real_image_probing(monkeypatch):
    """These tests exercise rules/mapping/export, not the image pipeline
    (that's test_image.py) -- stub probe_images so nothing here makes a
    real network call. Tests that DO want image integration override
    this via monkeypatch in the test body."""
    monkeypatch.setattr(image, "probe_images", lambda conn, urls, **kw: {})


def _insert_product(conn, sku, **overrides):
    base = {
        "sku": sku,
        "title": "A perfectly valid product title",
        "description": "A different description than the title",
        "image_link": None,
        "price_kes": 1500.0,
        "sale_price_kes": None,
        "brand_raw": "UGREEN",
        "product_type_raw": "Cables",
        "availability": "in stock",
        "condition": "new",
        "fetched_at": "2026-01-01T00:00:00Z",
        "feed_hash": f"hash-{sku}",
    }
    base.update(overrides)
    conn.execute(
        """INSERT INTO products (sku, title, description, image_link, price_kes, sale_price_kes,
                                  brand_raw, product_type_raw, availability, condition, fetched_at, feed_hash)
           VALUES (:sku, :title, :description, :image_link, :price_kes, :sale_price_kes,
                   :brand_raw, :product_type_raw, :availability, :condition, :fetched_at, :feed_hash)""",
        base,
    )
    conn.commit()


def _insert_resolution(conn, kind, raw_value, jumia_id, jumia_label):
    conn.execute(
        """INSERT INTO resolutions (kind, raw_value, jumia_id, jumia_label, confirmed_by_human, updated_at)
           VALUES (?, ?, ?, ?, 1, '2026-01-01T00:00:00Z')""",
        (kind, raw_value, jumia_id, jumia_label),
    )
    conn.commit()


def test_run_validation_marks_completed_and_counts_correctly(conn):
    _insert_resolution(conn, "brand", "UGREEN", "1118344", "Ugreen")
    _insert_resolution(conn, "category", "Cables", "1000473", "Computing / Cables")
    _insert_product(conn, "A1")
    _insert_product(conn, "A2", title="Too short")  # fails name_length

    result = pipeline.run_validation(conn, rules_path=str(REAL_RULES_PATH))

    assert result.total == 2
    assert result.passed == 1
    assert result.blocked == 1

    status = conn.execute("SELECT status FROM validation_runs WHERE id = ?", (result.run_id,)).fetchone()[0]
    assert status == "completed"


def test_run_validation_blocks_unresolved_brand(conn):
    _insert_resolution(conn, "category", "Cables", "1000473", "Computing / Cables")
    _insert_product(conn, "A1")  # brand never resolved

    result = pipeline.run_validation(conn, rules_path=str(REAL_RULES_PATH))
    assert result.blocked == 1

    rule_ids = {
        r[0] for r in conn.execute(
            "SELECT rule_id FROM row_issues WHERE run_id = ? AND sku = 'A1'", (result.run_id,)
        )
    }
    assert "brand_format" in rule_ids


def test_run_validation_wires_image_probe_results_into_rules(conn, monkeypatch):
    """Integration check: run_validation actually calls image.probe_images
    and threads its result into the rule engine, so an undersized image
    blocks the row via image_min_dims."""
    _insert_resolution(conn, "brand", "UGREEN", "1118344", "Ugreen")
    _insert_resolution(conn, "category", "Cables", "1000473", "Computing / Cables")
    url = "https://example.com/tiny.png"
    _insert_product(conn, "A1", image_link=url)

    tiny_image = image.ImageInfo(
        url=url, status_code=200, width=50, height=50, bytes=10,
        corner_luminance=250.0, checked_at="2026-01-01T00:00:00Z",
    )
    monkeypatch.setattr(image, "probe_images", lambda conn, urls, **kw: {url: tiny_image})

    result = pipeline.run_validation(conn, rules_path=str(REAL_RULES_PATH))

    rule_ids = {
        r[0] for r in conn.execute(
            "SELECT rule_id FROM row_issues WHERE run_id = ? AND sku = 'A1'", (result.run_id,)
        )
    }
    assert "image_min_dims" in rule_ids


def test_run_validation_persists_row_issues(conn):
    _insert_product(conn, "A2", title="Too short")

    result = pipeline.run_validation(conn, rules_path=str(REAL_RULES_PATH))

    issues = conn.execute(
        "SELECT sku, rule_id, severity FROM row_issues WHERE run_id = ?", (result.run_id,)
    ).fetchall()
    assert ("A2", "name_length", "block") in issues


def test_run_validation_marks_run_failed_on_exception(conn):
    _insert_product(conn, "A1")

    with pytest.raises(FileNotFoundError):
        pipeline.run_validation(conn, rules_path="/no/such/rules.yaml")

    row = conn.execute("SELECT status, error_message FROM validation_runs ORDER BY id DESC LIMIT 1").fetchone()
    assert row[0] == "failed"
    assert row[1] is not None


def test_run_export_raises_when_no_completed_run(conn):
    with pytest.raises(ValueError):
        pipeline.run_export(conn)


def test_run_export_only_includes_passed_and_warned_rows(tmp_path, conn, monkeypatch):
    _insert_resolution(conn, "brand", "UGREEN", "1118344", "Ugreen")
    _insert_resolution(conn, "category", "Cables", "1000473", "Computing / Cables")
    _insert_product(conn, "A1")
    _insert_product(conn, "A2", title="Too short")

    result = pipeline.run_validation(conn, rules_path=str(REAL_RULES_PATH))

    import openpyxl

    template_path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["SellerSKU", "Name", "Brand", "PrimaryCategory", "Price_KES", "Stock"])
    wb.save(template_path)

    monkeypatch.setattr(config, "UPLOAD_TEMPLATE_PATH", str(template_path))
    export_result = pipeline.run_export(conn, run_id=result.run_id, out_dir=str(tmp_path / "out"))

    assert export_result.rows_written == 1
    assert export_result.rows_rejected == 1

    wb_out = openpyxl.load_workbook(export_result.output_path)
    skus = [row[0].value for row in wb_out.active.iter_rows(min_row=2)]
    assert skus == ["A1"]
