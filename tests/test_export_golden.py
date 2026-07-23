"""Golden-file test: real fixture feed -> real rules -> exact expected export.

Protects against the single failure mode that causes mass Jumia rejection:
export column-order drift (Readme.md #15 principle 6). The template header
below is the real Upload_Template.xlsx header, confirmed 2026-07-23
against the actual file (Readme.md #13 Open Decision 5) -- kept as a
synthetic fixture here since the real file has real business data and
never belongs in git.
"""

import sqlite3
from pathlib import Path

import openpyxl

from jumia_feed_sync import config, db, image, ingest, pipeline

FEED_FIXTURE = Path(__file__).parent / "fixtures" / "sample_feed.xml"
REAL_RULES_PATH = Path(__file__).parent.parent / "config" / "rules.yaml"

TEMPLATE_HEADER = [
    "Name", "Name_AR", "Name_FR", "Description", "Description_AR", "Description_FR",
    "SellerSKU", "ParentSKU", "Brand", "PrimaryCategory", "AdditionalCategory", "GTIN_Barcode",
    "Price_KES", "Sale_Price_KES", "Sale_Price_Start_At", "Sale_Price_End_At", "Stock",
    "battery_capacity", "connection_gender", "cpu_manufacturer", "graphics_memory",
    "memory_technology", "panel_type", "processor_type", "storage_capacity", "variation",
    "certifications", "color", "color_AR", "color_FR", "color_family", "display_resolution",
    "display_size", "hdd_size", "main_material", "manufacturer_txt", "material_family",
    "memory_capacity", "model", "modem_type", "mount_type", "note", "package_content",
    "package_content_AR", "package_content_FR", "plug_type", "product_line", "product_measures",
    "product_warranty", "product_weight", "production_country", "short_description",
    "short_description_AR", "short_description_FR", "system_memory", "voltage",
    "warranty_address", "warranty_duration", "warranty_type", "youtube_id",
    "MainImage", "Image2", "Image3", "Image4", "Image5", "Image6", "Image7", "Image8",
]


def test_golden_end_to_end_export(tmp_path, monkeypatch):
    assert len(TEMPLATE_HEADER) == 68

    conn = sqlite3.connect(":memory:")
    db.migrate(conn)

    items = ingest.parse_feed(FEED_FIXTURE.read_bytes())
    ingest.upsert_products(conn, items)

    # Resolve exactly one item's brand+category -- UG-55551B -- leaving the
    # other four unresolved, to exercise both outcomes in one golden run.
    conn.execute(
        """INSERT INTO resolutions (kind, raw_value, jumia_id, jumia_label, confirmed_by_human, updated_at)
           VALUES ('brand', 'UGREEN', '1118344', 'Ugreen', 1, '2026-01-01T00:00:00Z')"""
    )
    conn.execute(
        """INSERT INTO resolutions (kind, raw_value, jumia_id, jumia_label, confirmed_by_human, updated_at)
           VALUES ('category', 'Components & Accessories', '1000473',
                   'Computing / Computer Accessories / Cables & Interconnects', 1, '2026-01-01T00:00:00Z')"""
    )
    conn.commit()

    template_path = tmp_path / "Upload_Template.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(TEMPLATE_HEADER)
    wb.save(template_path)
    monkeypatch.setattr(config, "UPLOAD_TEMPLATE_PATH", str(template_path))

    # Real image checks are exercised here too, with a mocked probe result
    # (no real network in tests) -- a conforming image, so it doesn't
    # change the pass/fail counts already driven by the other rules. See
    # test_golden_export_blocks_on_undersized_image below for the case
    # where the image check is what flips the outcome.
    ugreen_image_url = "https://res.cloudinary.com/dyfj4bj3c/image/upload/v1778347237/isgnw2vfpae2svwdqcfj.png"
    monkeypatch.setattr(
        image, "probe_images",
        lambda conn, urls, **kw: {
            ugreen_image_url: image.ImageInfo(
                url=ugreen_image_url, status_code=200, width=800, height=800, bytes=50000, format="PNG",
                corner_luminance=250.0, checked_at="2026-01-01T00:00:00Z",
            )
        },
    )

    validation = pipeline.run_validation(conn, rules_path=str(REAL_RULES_PATH))
    assert validation.total == 5
    assert validation.passed == 1
    assert validation.blocked == 4

    passing_row_issues = {
        r[0] for r in conn.execute(
            "SELECT rule_id FROM row_issues WHERE run_id = ? AND sku = 'UG-55551B'", (validation.run_id,)
        )
    }
    assert not passing_row_issues & {"image_reachable", "image_min_dims", "image_white_bg"}

    results = pipeline.run_export(conn, run_id=validation.run_id, out_dir=str(tmp_path / "out"))
    # One category (UG-55551B's) among the approved rows -> one file (Readme.md #11).
    assert len(results) == 1
    result = results[0]
    assert result.category == "1000473"
    assert result.rows_written == 1
    # rejects.csv is one row per (sku, rule) block violation, not per sku --
    # a blocked sku can fail several rules at once (Readme.md #11).
    assert result.rows_rejected > 4

    wb_out = openpyxl.load_workbook(result.output_path)
    ws = wb_out.active
    assert [c.value for c in ws[1]] == TEMPLATE_HEADER
    assert ws.max_row == 2

    exported = dict(zip(TEMPLATE_HEADER, [c.value for c in ws[2]]))
    assert exported["SellerSKU"] == "UG-55551B"
    assert exported["Name"] == "UGREEN GaN 500W Desktop Fast Charger (6-Port) UK - X759"
    assert exported["Brand"] == "1118344 - Ugreen"
    assert exported["PrimaryCategory"] == "1000473 - Computing / Computer Accessories / Cables & Interconnects"
    assert exported["Price_KES"] == 32000.0
    assert exported["Stock"] == config.STOCK_DEFAULT
    assert exported["MainImage"] == (
        "https://res.cloudinary.com/dyfj4bj3c/image/upload/v1778347237/isgnw2vfpae2svwdqcfj.png"
    )

    rejects_content = Path(result.rejects_path).read_text(encoding="utf-8")
    rejected_skus = {line.split(",")[0] for line in rejects_content.splitlines()[1:]}
    assert rejected_skus == {"POWER CAB 3 PIN", "981-000870", "SDSDXEP-064G-GN4IN", "12XD005LUM"}


def test_golden_export_blocks_on_undersized_image(monkeypatch):
    """Same setup as the golden test above, but the previously-passing
    row's image now fails image_min_dims -- proving the image check is
    actually wired into the export decision, not just present and inert."""
    conn = sqlite3.connect(":memory:")
    db.migrate(conn)

    items = ingest.parse_feed(FEED_FIXTURE.read_bytes())
    ingest.upsert_products(conn, items)
    conn.execute(
        """INSERT INTO resolutions (kind, raw_value, jumia_id, jumia_label, confirmed_by_human, updated_at)
           VALUES ('brand', 'UGREEN', '1118344', 'Ugreen', 1, '2026-01-01T00:00:00Z')"""
    )
    conn.execute(
        """INSERT INTO resolutions (kind, raw_value, jumia_id, jumia_label, confirmed_by_human, updated_at)
           VALUES ('category', 'Components & Accessories', '1000473',
                   'Computing / Computer Accessories / Cables & Interconnects', 1, '2026-01-01T00:00:00Z')"""
    )
    conn.commit()

    ugreen_image_url = "https://res.cloudinary.com/dyfj4bj3c/image/upload/v1778347237/isgnw2vfpae2svwdqcfj.png"
    monkeypatch.setattr(
        image, "probe_images",
        lambda conn, urls, **kw: {
            ugreen_image_url: image.ImageInfo(
                url=ugreen_image_url, status_code=200, width=50, height=50, bytes=500, format="PNG",
                corner_luminance=250.0, checked_at="2026-01-01T00:00:00Z",
            )
        },
    )

    validation = pipeline.run_validation(conn, rules_path=str(REAL_RULES_PATH))
    assert validation.passed == 0
    assert validation.blocked == 5

    rule_ids = {
        r[0] for r in conn.execute(
            "SELECT rule_id FROM row_issues WHERE run_id = ? AND sku = 'UG-55551B'", (validation.run_id,)
        )
    }
    assert "image_min_dims" in rule_ids
