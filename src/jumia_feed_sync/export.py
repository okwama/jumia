"""Export writer: approved rows -> a copy of Upload_Template.xlsx. See Readme.md #11.

Column order is derived from the template's own header row at runtime,
never hardcoded -- Jumia can change the template without notice
(Readme.md #14).
"""

from __future__ import annotations

import csv
import shutil

import openpyxl

from jumia_feed_sync.rules import Issue


def write_export(rows: list[dict], template_path: str, output_path: str) -> int:
    """Copies the template, clears existing data rows (2+), appends `rows`
    in the template's own header column order. Returns rows written."""
    shutil.copyfile(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.worksheets[0]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row in rows:
        ws.append([row.get(col) for col in header])

    wb.save(output_path)
    return len(rows)


def write_rejects_csv(issues: list[Issue], path: str) -> int:
    blocked = [i for i in issues if i.severity == "block"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["sku", "rule_id", "field", "message"])
        for issue in blocked:
            writer.writerow([issue.sku, issue.rule_id, issue.field or "", issue.message])
    return len(blocked)
