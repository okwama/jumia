"""Bootstrap harvesters for id_label_catalog. See Readme.md #6, #7.

Two independent sources feed the same catalog:

- `harvest()` -- a filled Upload_Template.xlsx. Only captures categories/
  brands actually in use, but confirms real current practice.
- `harvest_guidelines()` -- Jumia's own seller guidelines workbook
  (Brands/Categories sheets). The authoritative, complete reference:
  every valid brand and category code, straight from Jumia, with the
  correct numeric IDs (unlike the commission sheet's UUID SIDs).

ParentSKU is not an ID/label pair and real data shows it isn't reliably
derivable from the SKU string either (Readme.md #13 Open Decision 4), so
it's out of scope for both harvesters.
"""

from __future__ import annotations

import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone

import openpyxl

_ID_LABEL_RE = re.compile(r"^\s*(\d+)\s*-\s*(.+?)\s*$")

_TEMPLATE_COLUMN_KINDS = {
    "Brand": "brand",
    "PrimaryCategory": "category",
    "AdditionalCategory": "category",
}

_GUIDELINES_SHEET_KINDS = {
    "Brands": "brand",
    "Categories": "category",
}


@dataclass
class HarvestSummary:
    rows_scanned: int
    pairs_found: int
    pairs_new: int


def _header_index(header_row: tuple) -> dict[str, int]:
    return {str(value).strip(): idx for idx, value in enumerate(header_row) if value}


def harvest_from_workbook(path: str) -> tuple[int, list[tuple[str, str, str]]]:
    """Filled template -> (data row count, [(kind, jumia_id, jumia_label), ...])."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = _header_index(next(rows))
    columns = {name: header[name] for name in _TEMPLATE_COLUMN_KINDS if name in header}

    row_count = 0
    pairs: list[tuple[str, str, str]] = []
    for row in rows:
        if row is None or all(v is None for v in row):
            continue
        row_count += 1
        for column_name, col_idx in columns.items():
            raw = row[col_idx]
            if not raw:
                continue
            match = _ID_LABEL_RE.match(str(raw))
            if not match:
                continue
            pairs.append((_TEMPLATE_COLUMN_KINDS[column_name], match.group(1), match.group(2)))
    return row_count, pairs


def harvest_from_guidelines(path: str) -> tuple[int, list[tuple[str, str, str]]]:
    """Jumia guidelines workbook -> (data row count, [(kind, jumia_id, jumia_label), ...]).

    Brands/Categories sheets are single-column "ID - Label" lists, one
    entry per row, no header-position lookup needed beyond skipping row 1.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    row_count = 0
    pairs: list[tuple[str, str, str]] = []
    for sheet_name, kind in _GUIDELINES_SHEET_KINDS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = row[0] if row else None
            if not raw:
                continue
            row_count += 1
            match = _ID_LABEL_RE.match(str(raw))
            if not match:
                continue
            pairs.append((kind, match.group(1), match.group(2)))
    return row_count, pairs


def upsert_catalog(
    conn: sqlite3.Connection, pairs: list[tuple[str, str, str]], source: str = "template"
) -> tuple[int, int]:
    """Returns (unique pairs found, pairs new to the catalog)."""
    seen_at = datetime.now(timezone.utc).isoformat()
    existing = {
        (kind, jumia_id) for kind, jumia_id in conn.execute("SELECT kind, jumia_id FROM id_label_catalog")
    }

    unique_pairs = {(kind, jumia_id): jumia_label for kind, jumia_id, jumia_label in pairs}
    new = sum(1 for key in unique_pairs if key not in existing)

    for (kind, jumia_id), jumia_label in unique_pairs.items():
        conn.execute(
            """
            INSERT INTO id_label_catalog (kind, jumia_id, jumia_label, source, first_seen_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(kind, jumia_id) DO NOTHING
            """,
            (kind, jumia_id, jumia_label, source, seen_at),
        )
    conn.commit()
    return len(unique_pairs), new


def harvest(conn: sqlite3.Connection, path: str) -> HarvestSummary:
    row_count, pairs = harvest_from_workbook(path)
    pairs_found, pairs_new = upsert_catalog(conn, pairs, source="template")
    return HarvestSummary(rows_scanned=row_count, pairs_found=pairs_found, pairs_new=pairs_new)


def harvest_guidelines(conn: sqlite3.Connection, path: str) -> HarvestSummary:
    row_count, pairs = harvest_from_guidelines(path)
    pairs_found, pairs_new = upsert_catalog(conn, pairs, source="jumia_reference")
    return HarvestSummary(rows_scanned=row_count, pairs_found=pairs_found, pairs_new=pairs_new)
